#!/usr/bin/env python3

import base64
import countrycode
import csv
import dataclasses
import datetime
from functools import cache
import json
import logging
import openpyxl
import time


# header font
HFONT_STYLE = openpyxl.styles.Font(bold=True)

# border line styles
SIDE_BLACK_THIN = openpyxl.styles.borders.Side(style='thin', color='000000')
SIDE_BLACK_DOUBLE = openpyxl.styles.borders.Side(style='double', color='000000')

# cell border styles
HBORDER_STYLE = openpyxl.styles.borders.Border(top=SIDE_BLACK_THIN, bottom=SIDE_BLACK_DOUBLE, left=SIDE_BLACK_THIN, right=SIDE_BLACK_THIN)
BORDER_STYLE = openpyxl.styles.borders.Border(bottom=SIDE_BLACK_THIN, left=SIDE_BLACK_THIN, right=SIDE_BLACK_THIN)
MBORDER_STYLE = openpyxl.styles.borders.Border(top=SIDE_BLACK_THIN, bottom=SIDE_BLACK_THIN, left=SIDE_BLACK_THIN, right=SIDE_BLACK_THIN)

# fill styles: certificate type
FILL_CERT_TYPE_ROOT = openpyxl.styles.PatternFill(patternType='solid', fgColor='fcf3d0', bgColor='fcf3d0')
FILL_CERT_TYPE_INTERMEDIATE = openpyxl.styles.PatternFill(patternType='solid', fgColor='dceaf6', bgColor='dceaf6')

# fill style: revoked certificates
FILL_REVOKED = openpyxl.styles.PatternFill(patternType='solid', fgColor='ff3333', bgColor='ff3333')

# fill style: technically-constrained certificates
FILL_TECHNICALLY_CONSTRAINED = openpyxl.styles.PatternFill(patternType='solid', fgColor='e9f3ec', bgColor='e9f3ec')

# fill style: certificates not included in any root store
FILL_NOT_TRUSTED = openpyxl.styles.PatternFill(patternType='solid', fgColor='c0c0c0', bgColor='c0c0c0')


# logging stuff
logging.basicConfig(
    level=logging.INFO,
    format=r'%(asctime)s.%(msecs)03dZ [%(levelname)s] %(name)s: %(message)s',
    datefmt='%Y-%m-%dT%H:%M:%S'
)
logging.Formatter.converter = time.gmtime
Logger = logging.getLogger(__name__)


@cache
def get_country_code(country_name):
    candidate = None
    if len(country_name) == 2 and country_name.isascii():
        return country_name
    else:
        candidate = countrycode.countrycode(country_name, origin='country.name', destination='iso2c')
        if candidate:
            return candidate
        else:
            for fromcode in (e for e in countrycode.codelist.keys() if e.startswith('cldr.short.')):
                for pos, candidate_name in enumerate(countrycode.codelist[fromcode]):
                    if candidate_name.lower() == country_name.lower():
                        return countrycode.codelist['iso2c'][pos]
    return ''


def canonicalize(row):
    # Technically Constrained
    row[19] = (row[19].upper() == 'TRUE')
    # Audits Same as Parent?
    row[24] = (row[24].upper() == 'TRUE')
    # CP Same as Parent?
    row[62] = (row[62].upper() == 'TRUE')
    # CPS Same as Parent?
    row[65] = (row[65].upper() == 'TRUE')
    # CP/CPS Same as Parent?
    row[68] = (row[68].upper() == 'TRUE')
    # TLS Capable
    row[74] = (row[74].upper() == 'TRUE')
    # TLS EV Capable
    row[75] = (row[75].upper() == 'TRUE')
    # Code Signing Capable
    row[76] = (row[76].upper() == 'TRUE')
    # S/MIME Capable
    row[77] = (row[77].upper() == 'TRUE')

    # Authority Key Identifier
    if row[17] != '':
        row[17] = base64.b64decode(row[17]).hex(':')
    # Subject Key Identifier
    if row[18] != '':
        row[18] = base64.b64decode(row[18]).hex(':')

    # Valid From (GMT)
    row[15] = row[15].replace('.', '-')
    # Valid To (GMT)
    row[16] = row[16].replace('.', '-')
    # Standard Audit Statement Date
    row[27] = row[27].replace('.', '-')
    # Standard Audit Period Start Date
    row[28] = row[28].replace('.', '-')
    # Standard Audit Period End Date
    row[29] = row[29].replace('.', '-')
    # NetSec Audit Statement Date
    row[32] = row[32].replace('.', '-')
    # NetSec Audit Period Start Date
    row[33] = row[33].replace('.', '-')
    # NetSec Audit Period End Date
    row[34] = row[34].replace('.', '-')
    # TLS BR Audit Statement Date
    row[37] = row[37].replace('.', '-')
    # TLS BR Audit Period Start Date
    row[38] = row[38].replace('.', '-')
    # TLS BR Audit Period End Date
    row[39] = row[39].replace('.', '-')
    # TLS EVG Audit Statement Date
    row[42] = row[42].replace('.', '-')
    # TLS EVG Audit Period Start Date
    row[43] = row[43].replace('.', '-')
    # TLS EVG Audit Period End Date
    row[44] = row[44].replace('.', '-')
    # Code Signing Audit Statement Date
    row[47] = row[47].replace('.', '-')
    # Code Signing Audit Period Start Date
    row[48] = row[48].replace('.', '-')
    # Code Signing Audit Period End Date
    row[49] = row[49].replace('.', '-')
    # S/MIME BR Audit Statement Date
    row[52] = row[52].replace('.', '-')
    # S/MIME BR Audit Period Start Date
    row[53] = row[53].replace('.', '-')
    # S/MIME BR Audit Period End Date
    row[54] = row[54].replace('.', '-')
    # VMC Audit Statement Date
    row[57] = row[57].replace('.', '-')
    # VMC Audit Period Start Date
    row[58] = row[58].replace('.', '-')
    # VMC Audit Period End Date
    row[59] = row[59].replace('.', '-')
    # CP Last Update Date
    row[64] = row[64].replace('.', '-')
    # CPS Last Update Date
    row[67] = row[67].replace('.', '-')
    # CP/CPS Last Update Date
    row[70] = row[70].replace('.', '-')

    # JSON array
    if row[22] != '':
        row[22] = '\n'.join(json.loads(row[22]))


def add_metadata_sheet(metadata_sheet):
    metadata_sheet.column_dimensions['A'].width = 16
    metadata_sheet.column_dimensions['B'].width = 48

    row = [
        'source',
        'https://www.ccadb.org/resources',
    ]
    row = [openpyxl.cell.WriteOnlyCell(metadata_sheet, value=c) for c in row]
    for cell in row:
        cell.number_format = openpyxl.styles.numbers.FORMAT_TEXT
        cell.border = MBORDER_STYLE
    metadata_sheet.append(row)

    row = [
        'generated at',
        datetime.datetime.now(tz=datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
    ]
    row = [openpyxl.cell.WriteOnlyCell(metadata_sheet, value=c) for c in row]
    for cell in row:
        cell.number_format = openpyxl.styles.numbers.FORMAT_TEXT
        cell.border = MBORDER_STYLE
    metadata_sheet.append(row)

    row = [
        'generator',
        'https://github.com/kidmin/ccadb-certificates-tabular',
    ]
    row = [openpyxl.cell.WriteOnlyCell(metadata_sheet, value=c) for c in row]
    for cell in row:
        cell.number_format = openpyxl.styles.numbers.FORMAT_TEXT
        cell.border = MBORDER_STYLE
    metadata_sheet.append(row)


@dataclasses.dataclass(eq=False)
class CACertificate:
    is_root: bool
    is_trusted: bool | None
    children: set = dataclasses.field(default_factory=set)


def main():
    cert_by_sfid = dict()
    cert_forest_edges = list()
    cert_forest_roots = set()

    Logger.info('START pass 1 of loading the table')
    num_records = 0
    with open('AllCertificateRecordsCSVFormatv2', 'r', encoding='UTF-8', newline='') as csv_fh:
        csv_reader = csv.reader(csv_fh, dialect='unix')
        for row in csv_reader:
            num_records += 1

            sfid = row[1]
            paernt_sfid = row[3]
            is_root = row[5] == 'Root Certificate'
            if is_root:
                is_trusted = any(e.capitalize() == 'Included' for e in row[7:11])
            else:
                is_trusted = None
            cert_by_sfid[sfid] = CACertificate(is_root=is_root, is_trusted=is_trusted)
            cert_forest_edges.append((paernt_sfid, sfid))
    Logger.info('END pass 1 of loading the table')

    Logger.info('START building CA tree')
    for parent_sfid, child_sfid in cert_forest_edges:
        if parent_sfid in cert_by_sfid:
            cert_by_sfid[parent_sfid].children.add(child_sfid)
        else:
            cert_forest_roots.add(child_sfid)
    del cert_forest_edges
    hier_stack = [list(cert_forest_roots)]
    while hier_stack:
        if len(hier_stack) > 7:
            raise RuntimeError('loop detected while CA tree walk')
        if hier_stack[-1]:
            sfid = hier_stack[-1][-1]
            next_children = list(cert_by_sfid[sfid].children)
            for child_sfid in next_children:
                if not cert_by_sfid[child_sfid].is_root:
                    cert_by_sfid[child_sfid].is_trusted = cert_by_sfid[sfid].is_trusted
            hier_stack[-1].pop()
            if next_children:
                hier_stack.append(next_children)
        else:
            hier_stack.pop()
    del cert_forest_roots
    Logger.info('END building CA tree')

    Logger.info('START preparing workbook')
    book = openpyxl.Workbook(write_only=True)

    sheet = book.create_sheet(title='AllCertificateRecords')

    sheet.auto_filter.ref = f"A1:CH{num_records}"
    sheet.freeze_panes = 'D2'
    sheet.column_dimensions['A'].width = 14
    sheet.column_dimensions['B'].width = 4
    sheet.column_dimensions['C'].width = 36
    sheet.column_dimensions['D'].width = 4
    sheet.column_dimensions['E'].width = 24
    sheet.column_dimensions['F'].width = 22
    sheet.column_dimensions['G'].width = 24
    sheet.column_dimensions['H'].width = 18
    sheet.column_dimensions['I'].width = 18
    sheet.column_dimensions['J'].width = 18
    sheet.column_dimensions['K'].width = 18
    sheet.column_dimensions['L'].width = 4
    sheet.column_dimensions['M'].width = 8
    sheet.column_dimensions['N'].width = 8
    sheet.column_dimensions['O'].width = 16
    sheet.column_dimensions['P'].width = 4
    sheet.column_dimensions['Q'].width = 4
    sheet.column_dimensions['R'].width = 12
    sheet.column_dimensions['S'].width = 12
    sheet.column_dimensions['T'].width = 4
    sheet.column_dimensions['U'].width = 4
    sheet.column_dimensions['V'].width = 8
    sheet.column_dimensions['W'].width = 36
    sheet.column_dimensions['X'].width = 14
    sheet.column_dimensions['Y'].width = 14
    sheet.column_dimensions['Z'].width = 8
    sheet.column_dimensions['AA'].width = 24
    sheet.column_dimensions['AB'].width = 8
    sheet.column_dimensions['AC'].width = 14
    sheet.column_dimensions['AD'].width = 14
    sheet.column_dimensions['AE'].width = 12
    sheet.column_dimensions['AF'].width = 12
    sheet.column_dimensions['AG'].width = 12
    sheet.column_dimensions['AH'].width = 14
    sheet.column_dimensions['AI'].width = 14
    sheet.column_dimensions['AJ'].width = 12
    sheet.column_dimensions['AK'].width = 12
    sheet.column_dimensions['AL'].width = 12
    sheet.column_dimensions['AM'].width = 14
    sheet.column_dimensions['AN'].width = 14
    sheet.column_dimensions['AO'].width = 12
    sheet.column_dimensions['AP'].width = 12
    sheet.column_dimensions['AQ'].width = 12
    sheet.column_dimensions['AR'].width = 14
    sheet.column_dimensions['AS'].width = 14
    sheet.column_dimensions['AT'].width = 12
    sheet.column_dimensions['AU'].width = 12
    sheet.column_dimensions['AV'].width = 12
    sheet.column_dimensions['AW'].width = 14
    sheet.column_dimensions['AX'].width = 14
    sheet.column_dimensions['AY'].width = 12
    sheet.column_dimensions['AZ'].width = 12
    sheet.column_dimensions['BA'].width = 12
    sheet.column_dimensions['BB'].width = 14
    sheet.column_dimensions['BC'].width = 14
    sheet.column_dimensions['BD'].width = 12
    sheet.column_dimensions['BE'].width = 12
    sheet.column_dimensions['BF'].width = 12
    sheet.column_dimensions['BG'].width = 14
    sheet.column_dimensions['BH'].width = 14
    sheet.column_dimensions['BI'].width = 12
    sheet.column_dimensions['BJ'].width = 12
    sheet.column_dimensions['BK'].width = 12
    sheet.column_dimensions['BL'].width = 14
    sheet.column_dimensions['BM'].width = 14
    sheet.column_dimensions['BN'].width = 8
    sheet.column_dimensions['BO'].width = 14
    sheet.column_dimensions['BP'].width = 12
    sheet.column_dimensions['BQ'].width = 8
    sheet.column_dimensions['BR'].width = 14
    sheet.column_dimensions['BS'].width = 12
    sheet.column_dimensions['BT'].width = 8
    sheet.column_dimensions['BU'].width = 14
    sheet.column_dimensions['BV'].width = 12
    sheet.column_dimensions['BW'].width = 14
    sheet.column_dimensions['BX'].width = 14
    sheet.column_dimensions['BY'].width = 14
    sheet.column_dimensions['BZ'].width = 8
    sheet.column_dimensions['CA'].width = 8
    sheet.column_dimensions['CB'].width = 8
    sheet.column_dimensions['CC'].width = 8
    sheet.column_dimensions['CD'].width = 14
    sheet.column_dimensions['CE'].width = 14
    sheet.column_dimensions['CF'].width = 12
    sheet.column_dimensions['CG'].width = 4
    sheet.column_dimensions['CH'].width = 4

    cert_type_rules = [
        openpyxl.formatting.rule.CellIsRule(
            operator='equal',
            formula=['"Root Certificate"'],
            stopIfTrue=False,
            fill=FILL_CERT_TYPE_ROOT
            ),
        openpyxl.formatting.rule.CellIsRule(
            operator='equal',
            formula=['"Intermediate Certificate"'],
            stopIfTrue=False,
            fill=FILL_CERT_TYPE_INTERMEDIATE
            ),
    ]
    for rule in cert_type_rules:
        sheet.conditional_formatting.add(f"F2:F{num_records}", rule)

    cert_revoked_rules = [
        openpyxl.formatting.rule.CellIsRule(
            operator='equal',
            formula=['"Revoked"'],
            stopIfTrue=False,
            fill=FILL_REVOKED
            ),
        openpyxl.formatting.rule.CellIsRule(
            operator='equal',
            formula=['"Parent Cert Revoked"'],
            stopIfTrue=False,
            fill=FILL_REVOKED
            ),
    ]
    for rule in cert_revoked_rules:
        sheet.conditional_formatting.add(f"O2:O{num_records}", rule)

    cert_constrained_rules = [
        openpyxl.formatting.rule.CellIsRule(
            operator='equal',
            formula=[True],
            stopIfTrue=False,
            fill=FILL_TECHNICALLY_CONSTRAINED
            ),
    ]
    for rule in cert_constrained_rules:
        sheet.conditional_formatting.add(f"V2:V{num_records}", rule)

    cert_not_trusted_rules = [
        openpyxl.formatting.rule.CellIsRule(
            operator='equal',
            formula=[False],
            stopIfTrue=False,
            fill=FILL_NOT_TRUSTED
            ),
    ]
    for rule in cert_not_trusted_rules:
        sheet.conditional_formatting.add(f"M2:M{num_records}", rule)
        sheet.conditional_formatting.add(f"N2:N{num_records}", rule)
    Logger.info('END preparing workbook')

    Logger.info('START pass 2 of loading the table')
    with open('AllCertificateRecordsCSVFormatv2', 'r', encoding='UTF-8', newline='') as csv_fh:
        csv_reader = csv.reader(csv_fh, dialect='unix')
        header = next(csv_reader)
        header.append('X-Country (alpha-2)')
        header.append('X-crt.sh link')
        header.insert(80, header.pop(78))
        header.insert(23, 'X-Number of items in "JSON Array of Partitioned CRLs"')
        header.insert(12, 'X-Chains up to Roots Included in any Root Store?')
        header.insert(12, 'X-Included in any Root Store?')
        header = [openpyxl.cell.WriteOnlyCell(sheet, value=hc) for hc in header]
        for hc in header:
            hc.font = HFONT_STYLE
            hc.border = HBORDER_STYLE
            hc.number_format = openpyxl.styles.numbers.FORMAT_TEXT
        sheet.row_dimensions[1].height = 14.25
        sheet.append(header)

        for row_no, row in enumerate(csv_reader, 2):
            if len(row) != 81:
                raise RuntimeError(f"unexpected number of rows {len(row)} at CSV line {row_no}")
            canonicalize(row)

            # X-Country (alpha-2)
            row.append(row.pop(78))
            row.append(get_country_code(row[80]))

            # X-crt.sh link
            row.append(f"https://crt.sh/?sha256={row[13]}")

            # X-Number of items in "JSON Array of Partitioned CRLs"
            if row[22] != '':
                row.insert(23, row[22].count('\n') + 1)
            else:
                row.insert(23, '')

            # X-Chains up to Roots Included in any Root Store?
            if row[5] == 'Intermediate Certificate':
                row.insert(12, cert_by_sfid[row[1]].is_trusted)
            else:
                row.insert(12, None)

            # X-Included in any Root Store?
            row.insert(12, any(e.capitalize() == 'Included' for e in row[7:11]))

            row = [openpyxl.cell.WriteOnlyCell(sheet, value=c) for c in row]
            for col_idx, cell in enumerate(row):
                cell.border = BORDER_STYLE
                if col_idx in {12, 13, 21, 27, 65, 68, 71, 77, 78, 79, 80}:
                    cell.number_format = openpyxl.styles.numbers.FORMAT_GENERAL
                elif col_idx in {17, 18, 30, 31, 32, 35, 36, 37, 40, 41, 42, 45, 46, 47, 50, 51, 52, 55, 56, 57, 60, 61, 62, 67, 70, 73}:
                    cell.number_format = openpyxl.styles.numbers.FORMAT_DATE_YYYYMMDD2
                    if cell.value != '':
                        cell.value = datetime.date.fromisoformat(cell.value)
                    else:
                        cell.value = None
                elif col_idx in {25}:
                    cell.number_format = openpyxl.styles.numbers.FORMAT_NUMBER
                elif col_idx in {82}:
                    cell.number_format = openpyxl.styles.numbers.FORMAT_TEXT
                elif col_idx in {85}:
                    cell.number_format = openpyxl.styles.numbers.FORMAT_TEXT
                    href = cell.value
                    cell.value = '\U0001F4DC'
                    cell.hyperlink = href
                else:
                    cell.number_format = openpyxl.styles.numbers.FORMAT_TEXT
            sheet.row_dimensions[row_no].height = 13.5
            sheet.append(row)
    Logger.info('END pass 2 of loading the table')

    add_metadata_sheet(book.create_sheet(title='_metadata'))

    book.active = book.worksheets[0]

    Logger.info('START saving workbook')
    book.save('CCADB-certificates.xlsx')
    Logger.info('END saving workbook')


if __name__ == '__main__':
    main()


# vim: set fileencoding=utf-8 nobomb fileformat=unix filetype=python number expandtab tabstop=8 softtabstop=4 shiftwidth=4 autoindent smartindent :
